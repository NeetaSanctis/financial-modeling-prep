import requests
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment
from openpyxl.utils import get_column_letter
import warnings
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import subprocess
import sys
import base64
import tempfile
import os
import datetime

warnings.simplefilter("ignore", FutureWarning)

BASE = "https://financialmodelingprep.com/api/v3"
API_KEY = "Enter Your API"
DEFAULT_OUTPUT_DIR = Path.home() / "Desktop" / "FullHistory"
CREATE_EXCEL = True

ENDPOINTS = [
    ("Annual_IncomeSheet",     "income-statement",        {}),
    ("Annual_BalanceSheet",    "balance-sheet-statement", {}),
    ("Annual_CashFlow",        "cash-flow-statement",     {}),
    ("Quarterly_IncomeSheet",  "income-statement",        {"period": "quarter"}),
    ("Quarterly_BalanceSheet", "balance-sheet-statement", {"period": "quarter"}),
    ("Quarterly_Cashflow",     "cash-flow-statement",     {"period": "quarter"}),
]

# ── Embedded icon ────────────────────────────────────────────────
_ICO_B64 = (
    "AAABAAEAEBAAAAAAIACUAAAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAAQAAAAEAgGAAAAH/P/YQAAAFtJREFUeJxjZGBgYBCr"
    "sPnPQCZgIkXzy6hvmAaQqhndEKIMQNeEzCfOBbpnUbjiy7hINAAPYMEn+bL9MGUGoACoN8QrbQkbQIzNMEBxGAwDAxgp"
    "yUhUcQHFBgAA+SkYi9hcQnQAAAAASUVORK5CYII="
)
_PNG_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAABdUlEQVR4nO2bsRWCMBiEo8/KAWQKK1tdxEl0ApzERbC1"
    "cgocwFYLH+8BBv4EIveH3NcpIe/uuABFMIaQpFn4DN6c9u9/CQnN83Jz8uY0KCbjbaQgltIEMZs3RtbfmU7sxm3Y2iA2"
    "QCvl8RVkHmsA2q9+Zd43BJuvnwBiMd/1W6LtL6ol0GV2zHJoBKD56ksmfUKo+1wNlzQNPsaqsdl17XyO6iUwtNo+56kN"
    "YOxjzvV8lQGEesa7zKMugFDmXedTdxOs38DKvDDmsRs1h4S6BkwNA0ALQMMA0ALQMAC0ADSTvweUeSGOyc6HCZR8YQPQ"
    "AkS2997DY9uSfAMYAFoAGgaAFoCGAaAFoAnyHqDt7c6H5BvAANAC0DAAtAA0DAAtAA0DQAtAwwDQAtAwALQANAwALQBN"
    "8gE0dk9r3igZmmrnePINaATg+plJ7NR9sgHtP+begrY/awPmGsKsPpkJhXil5/Bo7Gu02IDYl4OkP/kvR0nqfAAMqXAL"
    "WD1wvgAAAABJRU5ErkJggg=="
)

# ── Colour palette ───────────────────────────────────────────────
NAV_BG    = "#1a3a5c"   # dark navy header
ACCENT    = "#2a9d6e"   # green accent
ACCENT_HV = "#228b5e"   # green hover
PAGE_BG   = "#f0f2f5"   # light grey page
CARD_BG   = "#ffffff"   # white card
INPUT_BG  = "#f5f7fa"   # input field background
BORDER    = "#dde1e8"   # subtle border
TXT_MAIN  = "#1a1f2e"   # primary text
TXT_SUB   = "#6b7280"   # muted text
OK_FG     = "#1d6f42"   # success green text
ERR_FG    = "#b91c1c"   # error red text


def _write_temp_icon():
    data = base64.b64decode(_ICO_B64)
    tmp  = tempfile.NamedTemporaryFile(suffix=".ico", delete=False)
    tmp.write(data); tmp.close()
    return tmp.name

def _get_tk_icon_image():
    data = base64.b64decode(_PNG_B64)
    tmp  = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp.write(data); tmp.close()
    return tmp.name

def desktop_notify(title, message):
    try:
        if sys.platform == "win32":
            ps = (
                f"Add-Type -AssemblyName System.Windows.Forms;"
                f"$n=New-Object System.Windows.Forms.NotifyIcon;"
                f"$n.Icon=[System.Drawing.SystemIcons]::Information;"
                f"$n.Visible=$true;"
                f"$n.ShowBalloonTip(4000,'{title}','{message}',"
                f"[System.Windows.Forms.ToolTipIcon]::Info);"
                f"Start-Sleep -Seconds 5;$n.Dispose()"
            )
            subprocess.Popen(["powershell", "-WindowStyle", "Hidden", "-Command", ps],
                             creationflags=subprocess.CREATE_NO_WINDOW)
        elif sys.platform == "darwin":
            subprocess.Popen(["osascript", "-e",
                              f'display notification "{message}" with title "{title}"'])
        else:
            subprocess.Popen(["notify-send", title, message])
    except Exception:
        pass

def fetch_company_profile(session, ticker, api_key):
    r    = session.get(f"{BASE}/profile/{ticker}", params={"apikey": api_key}, timeout=60)
    r.raise_for_status()
    data = r.json()
    if isinstance(data, dict) and data.get("Error Message"):
        raise ValueError(data["Error Message"])
    if not isinstance(data, list) or not data:
        raise ValueError(f"No profile found for {ticker}")
    return data[0]

def normalize_ticker(raw: str):
    t = raw.strip().upper()
    if "." in t: return t, t.split(".")[0]
    if t.isdigit(): return t + ".T", t
    return t, t

def fetch_statement(session, endpoint, ticker, api_key, extra_params=None, limit=400):
    params = {"apikey": api_key, "limit": limit}
    if extra_params: params.update(extra_params)
    r    = session.get(f"{BASE}/{endpoint}/{ticker}", params=params, timeout=60)
    r.raise_for_status()
    data = r.json()
    if isinstance(data, dict) and data.get("Error Message"):
        raise ValueError(data["Error Message"])
    if not isinstance(data, list):
        raise ValueError(f"Unexpected response: {data}")
    return pd.DataFrame(data)

def sanitize_sheet_name(name): return name[:31]

def format_sheet(ws):
    nf = Font(bold=False); nb = Border()
    al = Alignment(horizontal="left", vertical="bottom")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = nf; cell.border = nb; cell.alignment = al
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

def build_file_for_ticker(raw_ticker, outdir):
    api_ticker, file_stem = normalize_ticker(raw_ticker)
    session = requests.Session()
    frames  = {}
    for sheet_name, endpoint, extra in ENDPOINTS:
        limit = 250 if extra.get("period") == "quarter" else 120
        frames[sheet_name] = fetch_statement(session, endpoint, api_ticker,
                                             API_KEY, extra_params=extra, limit=limit)
    profile    = fetch_company_profile(session, api_ticker, API_KEY)
    profile_df = pd.DataFrame(list(profile.items()), columns=["Field", "Value"])
    xlsx_path  = outdir / f"{file_stem}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sn, df in frames.items():
            df.T.to_excel(writer, sheet_name=sanitize_sheet_name(sn),
                          header=False, index=True, startrow=1)
        profile_df.to_excel(writer, sheet_name="Company_Info",
                            header=False, index=False)
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets: format_sheet(ws)
    wb.save(xlsx_path)

def open_folder(path: Path):
    if sys.platform == "win32":   subprocess.Popen(["explorer", str(path)])
    elif sys.platform == "darwin": subprocess.Popen(["open", str(path)])
    else:                          subprocess.Popen(["xdg-open", str(path)])

def save_log(outdir: Path, entries: list):
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = outdir / f"download_log_{ts}.csv"
    pd.DataFrame(entries, columns=["Ticker", "Status", "Message", "Timestamp"]
                 ).to_csv(path, index=False)
    return path


def get_tickers_from_dialog():
    result     = {"value": None}
    output_dir = {"path": DEFAULT_OUTPUT_DIR}

    root = tk.Tk()
    root.title("Ticker Download")
    root.geometry("620x600")
    root.resizable(False, False)
    root.configure(bg=PAGE_BG)

    # ── Icon ──────────────────────────────────────────────────────
    _ico_path = _write_temp_icon()
    _png_path = _get_tk_icon_image()
    try:
        if sys.platform == "win32":
            root.iconbitmap(_ico_path)
        else:
            root.iconphoto(True, tk.PhotoImage(file=_png_path))
    except Exception:
        pass

    # ── Helpers ───────────────────────────────────────────────────
    def lbl(parent, text, font=("Segoe UI", 9), fg=TXT_MAIN, **kw):
        w = tk.Label(parent, text=text, font=font, fg=fg,
                     bg=kw.pop("bg", PAGE_BG), **kw)
        return w

    def card(parent, **kw):
        return tk.Frame(parent, bg=CARD_BG,
                        highlightbackground=BORDER, highlightthickness=1, **kw)

    def btn(parent, text, command, bg=ACCENT, fg="#ffffff",
            font=("Segoe UI", 9, "bold"), width=14, **kw):
        b = tk.Button(parent, text=text, command=command,
                      bg=bg, fg=fg, font=font, width=width,
                      relief="flat", cursor="hand2",
                      activebackground=ACCENT_HV, activeforeground="#ffffff",
                      bd=0, padx=10, pady=6, **kw)
        b.bind("<Enter>", lambda e: b.config(bg=ACCENT_HV if bg == ACCENT else "#c8cdd6"))
        b.bind("<Leave>", lambda e: b.config(bg=bg))
        return b

    # ── NAV HEADER ────────────────────────────────────────────────
    nav = tk.Frame(root, bg=NAV_BG, height=52)
    nav.pack(fill="x")
    nav.pack_propagate(False)

    dot = tk.Frame(nav, bg=ACCENT, width=28, height=28)
    dot.place(x=18, rely=0.5, anchor="w")
    lbl(dot, "↓", font=("Segoe UI", 13, "bold"),
        fg="#ffffff", bg=ACCENT).place(relx=0.5, rely=0.5, anchor="center")

    lbl(nav, "  Ticker Download",
        font=("Segoe UI", 12, "bold"), fg="#ffffff", bg=NAV_BG).place(x=54, rely=0.5, anchor="w")

    # ── BODY ──────────────────────────────────────────────────────
    body = tk.Frame(root, bg=PAGE_BG)
    body.pack(fill="both", expand=True, padx=18, pady=14)

    # ── OUTPUT FOLDER CARD ────────────────────────────────────────
    fc = card(body)
    fc.pack(fill="x", pady=(0, 10))
    fi = tk.Frame(fc, bg=CARD_BG)
    fi.pack(fill="x", padx=14, pady=10)

    lbl(fi, "Output Folder", font=("Segoe UI", 9, "bold"),
        bg=CARD_BG).pack(anchor="w")

    folder_row = tk.Frame(fi, bg=CARD_BG)
    folder_row.pack(fill="x", pady=(5, 0))

    folder_var = tk.StringVar(value=str(DEFAULT_OUTPUT_DIR))
    folder_entry = tk.Entry(folder_row, textvariable=folder_var,
                            font=("Segoe UI", 9), bg=INPUT_BG,
                            fg=TXT_MAIN, relief="flat",
                            highlightbackground=BORDER, highlightthickness=1)
    folder_entry.pack(side="left", fill="x", expand=True, ipady=5)

    def browse():
        chosen = filedialog.askdirectory(title="Select Output Folder",
                                         initialdir=str(output_dir["path"]))
        if chosen:
            output_dir["path"] = Path(chosen)
            folder_var.set(chosen)

    btn(folder_row, "Browse", browse, width=8,
        font=("Segoe UI", 9)).pack(side="left", padx=(8, 0))

    # ── TICKER INPUT CARD ─────────────────────────────────────────
    tc = card(body)
    tc.pack(fill="x", pady=(0, 10))
    ti = tk.Frame(tc, bg=CARD_BG)
    ti.pack(fill="x", padx=14, pady=10)

    th = tk.Frame(ti, bg=CARD_BG)
    th.pack(fill="x")
    lbl(th, "Tickers", font=("Segoe UI", 9, "bold"), bg=CARD_BG).pack(side="left")

    def import_tickers():
        fp = filedialog.askopenfilename(
            title="Import Tickers",
            filetypes=[("Text/CSV", "*.txt *.csv"), ("All files", "*.*")])
        if not fp: return
        try:
            ext = Path(fp).suffix.lower()
            tickers = (pd.read_csv(fp, header=None).iloc[:, 0]
                       .dropna().astype(str).tolist()
                       if ext == ".csv"
                       else [l.strip() for l in open(fp) if l.strip()])
            text_box.config(state="normal")
            text_box.delete("1.0", "end")
            text_box.insert("1.0", "\n".join(tickers))
        except Exception as e:
            messagebox.showerror("Import Error", str(e))

    btn(th, "Import File", import_tickers,
        bg="#e8edf2", fg=TXT_MAIN,
        font=("Segoe UI", 8), width=10).pack(side="right")

    lbl(ti, "Separate by comma or newline  (e.g.  AAPL, MSFT, 7203)",
        font=("Segoe UI", 8), fg=TXT_SUB, bg=CARD_BG).pack(anchor="w", pady=(2, 5))

    text_box = tk.Text(ti, font=("Consolas", 10), height=8,
                       bg=INPUT_BG, fg=TXT_MAIN, relief="flat",
                       highlightbackground=BORDER, highlightthickness=1,
                       insertbackground=TXT_MAIN, padx=8, pady=6)
    text_box.pack(fill="x")

    # ── DOWNLOAD BUTTON ───────────────────────────────────────────
    dl_btn = btn(body, "  ↓  Download", None,
                 font=("Segoe UI", 10, "bold"), width=0)
    dl_btn.pack(fill="x", pady=(0, 10), ipady=4)

    # ── PROGRESS CARD ─────────────────────────────────────────────
    pc = card(body)
    pc.pack(fill="x", pady=(0, 10))
    pi = tk.Frame(pc, bg=CARD_BG)
    pi.pack(fill="x", padx=14, pady=10)

    prog_header = tk.Frame(pi, bg=CARD_BG)
    prog_header.pack(fill="x")
    lbl(prog_header, "Progress", font=("Segoe UI", 9, "bold"),
        bg=CARD_BG).pack(side="left")
    prog_count = lbl(prog_header, "", font=("Segoe UI", 9, "bold"),
                     fg=ACCENT, bg=CARD_BG)
    prog_count.pack(side="right")

    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Green.Horizontal.TProgressbar",
                    troughcolor=INPUT_BG, background=ACCENT,
                    darkcolor=ACCENT, lightcolor=ACCENT,
                    bordercolor=BORDER, thickness=8)

    bar = ttk.Progressbar(pi, orient="horizontal", mode="determinate",
                          style="Green.Horizontal.TProgressbar")
    bar.pack(fill="x", pady=(6, 4))

    prog_status = lbl(pi, "Waiting to start…",
                      font=("Segoe UI", 8), fg=TXT_SUB, bg=CARD_BG)
    prog_status.pack(anchor="w")

    # ── OUTPUT CARD ───────────────────────────────────────────────
    oc = card(body)
    oc.pack(fill="x")
    oi = tk.Frame(oc, bg=CARD_BG)
    oi.pack(fill="x", padx=14, pady=10)

    out_row = tk.Frame(oi, bg=CARD_BG)
    out_row.pack(fill="x")
    lbl(out_row, "Output", font=("Segoe UI", 9, "bold"),
        bg=CARD_BG).pack(side="left")
    open_btn = btn(out_row, "Open Folder", lambda: open_folder(output_dir["path"]),
                   bg="#e8edf2", fg=TXT_MAIN,
                   font=("Segoe UI", 8), width=10, state="disabled")
    open_btn.pack(side="right")

    out_lbl = lbl(oi, "", font=("Segoe UI", 8), fg=TXT_SUB,
                  bg=CARD_BG, justify="left", wraplength=560, anchor="w")
    out_lbl.pack(anchor="w", pady=(4, 0))

    # ── LOGIC ─────────────────────────────────────────────────────
    def submit():
        raw = text_box.get("1.0", "end").strip()
        if not raw: return
        tickers = [t.strip() for t in raw.replace("\n", ",").split(",") if t.strip()]
        if not tickers: return

        output_dir["path"] = Path(folder_var.get()) if folder_var.get() else DEFAULT_OUTPUT_DIR

        dl_btn.config(state="disabled")
        open_btn.config(state="disabled")
        text_box.config(state="disabled")
        out_lbl.config(text="", fg=TXT_SUB)

        total = len(tickers)
        bar["maximum"] = total
        bar["value"]   = 0
        prog_count.config(text=f"0 / {total}")
        prog_status.config(text="Starting…")
        root.update_idletasks()

        output_dir["path"].mkdir(parents=True, exist_ok=True)
        log_entries = []
        errors      = []

        for i, ticker in enumerate(tickers, 1):
            prog_status.config(text=f"Processing {ticker}…")
            prog_count.config(text=f"{i} / {total}")
            root.update_idletasks()
            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                build_file_for_ticker(ticker, output_dir["path"])
                log_entries.append((ticker, "SUCCESS", "", ts))
            except Exception as e:
                errors.append((ticker, str(e)))
                log_entries.append((ticker, "FAILED", str(e), ts))
            bar["value"] = i
            root.update_idletasks()

        prog_status.config(text="Done")
        log_path = save_log(output_dir["path"], log_entries)

        if errors:
            out_lbl.config(
                text="Completed with errors:\n"
                     + "\n".join(f"  •  {t}: {e}" for t, e in errors)
                     + f"\n\nLog saved: {log_path.name}",
                fg=ERR_FG)
            desktop_notify("Ticker Download",
                           f"{total-len(errors)}/{total} succeeded. {len(errors)} error(s).")
        else:
            out_lbl.config(
                text=f"✓  All {total} file(s) saved to:\n"
                     f"   {output_dir['path'].resolve()}\n"
                     f"   Log: {log_path.name}",
                fg=OK_FG)
            desktop_notify("Ticker Download",
                           f"All {total} ticker(s) downloaded successfully.")

        open_btn.config(state="normal")
        dl_btn.config(state="normal")
        text_box.config(state="normal")

        try: os.unlink(_ico_path); os.unlink(_png_path)
        except Exception: pass

    def on_close():
        try: os.unlink(_ico_path); os.unlink(_png_path)
        except Exception: pass
        root.destroy()

    dl_btn.config(command=submit)
    root.protocol("WM_DELETE_WINDOW", on_close)
    root.bind("<Control-Return>", lambda e: submit())
    root.mainloop()
    return result["value"]


def main():
    if API_KEY == "PUT_YOUR_NEW_FMP_KEY_HERE":
        raise ValueError("Please paste your FMP API key into API_KEY first.")
    get_tickers_from_dialog()

if __name__ == "__main__":
    main()
